import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


folder_path = './Generated_Files'
# folder_path = './files'
output_path = './output'
output_file = '_output.xlsx'
course_code =['CSPC601', 'CSPC602', 'CSPC603', 'CSPC604', 'CSPC605', 'CSPC606']
# course_code = []
marks = []
file_list = os.listdir(folder_path)

excel_files = [file for file in file_list if file.endswith('.xlsx') or file.endswith('.xls')]

if not os.path.exists(output_path):
    os.makedirs(output_path)
# for file in excel_files:
#     file_path = os.path.join(folder_path, file)
#     df = pd.read_excel(file_path)
#     course_code.append(df.iloc[7, 3])
#     marks.append(df.iloc[8, 3])
    
# print(course_code, marks) 
for i, file in enumerate(excel_files):
    file_path = os.path.join(folder_path, file)
    df = pd.read_excel(file_path)
    # df = pd.read_excel(file_path, skiprows=10, header=0)
    # print(marks[i])
    # mark_40 = marks[i]*0.4
    # mark_75 = marks[i]*0.75
    mark_40 = 16
    mark_75 = 30
    
    total_students = len(df)
    total_students_appeared = len(df[df['Marks'] != 'Absent'])
    total_absent = total_students - total_students_appeared
    avg_marks = df[df['Marks'] != 'Absent']['Marks'].astype(float).mean()
    less_than_16 = len(df[df['Marks'].astype(float) < mark_40])
    between_16_and_30 = len(df[(df['Marks'].astype(float) >= mark_40) & (df['Marks'].astype(float) < mark_75)])
    more_than_30 = len(df[df['Marks'].astype(float) >= mark_75])

    results = [{
        'Total Students': total_students,
        'Total Students Appeared': total_students_appeared,
        'Total Absent': total_absent,
        'Average Marks': avg_marks,
        'Students Less than 16': less_than_16,
        'Students Between 16 and 30': between_16_and_30,
        'Students More than 30': more_than_30
    }]
    results_df = pd.DataFrame(results)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Student Summary'


    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws1.append(r)


    mark_less_than = float(input(f"Enter a mark to filter students who scored less than that mark for file {file}: "))

    filtered_less_than_df = df[df['Marks'].astype(float) < mark_less_than]
    filtered_less_than_df = filtered_less_than_df.sort_values(by='Marks', ascending=True)
    if 'Unnamed: 4' in filtered_less_than_df:
        filtered_less_than_df= filtered_less_than_df.drop(['Unnamed: 4'], axis=1)
    ws2 = wb.create_sheet('Slow Learners')
    for r in dataframe_to_rows(filtered_less_than_df, index=False, header=True):
        ws2.append(r)

    mark_more_than = float(input(f"Enter a mark to filter students who scored more than that mark for file {file}: "))

    filtered_more_than_df = df[df['Marks'].astype(float) > mark_more_than]
    filtered_more_than_df = filtered_more_than_df.sort_values(by='Marks', ascending=False)
    if 'Unnamed: 4' in filtered_more_than_df:
        filtered_more_than_df= filtered_more_than_df.drop(['Unnamed: 4'], axis=1)
    ws3 = wb.create_sheet('Fast Learners')
    for r in dataframe_to_rows(filtered_more_than_df, index=False, header=True):
        ws3.append(r)


    output_file_path = os.path.join(output_path, file.replace('.xlsx', '') + output_file)
    wb.save(output_file_path)

    print("Results saved successfully to:", output_file_path)
    
#Task 2    
output_folder_path = './output'
output_files = [file for file in os.listdir(output_folder_path) if file.endswith('.xlsx')]

data_sheet1 = []
for idx, file in enumerate(output_files):
    file_path = os.path.join(output_folder_path, file)
    df = pd.read_excel(file_path, sheet_name='Student Summary', header=0)
    data_sheet1.append([course_code[idx]] + df.iloc[0].values.tolist())

columns_sheet1 = ['Course Code', 'Total Students', 'Total Students Appeared', 'Total Absent', 
                  'Average Marks', 'Students Less than 16', 
                  'Students Between 16 and 30', 'Students More than 30']

df_sheet1_transposed = pd.DataFrame(data_sheet1, columns=columns_sheet1).T.reset_index()
df_sheet1_transposed.columns = df_sheet1_transposed.iloc[0]
df_sheet1_transposed = df_sheet1_transposed.drop(0)

sheet2_roll_counts = {}
sheet3_roll_counts = {}

df_combined_sheet2 = pd.DataFrame()
df_combined_sheet3 = pd.DataFrame()

for file in output_files:
    file_path = os.path.join(output_folder_path, file)

    df_sheet2 = pd.read_excel(file_path, sheet_name='Slow Learners', header=0)
    df_sheet3 = pd.read_excel(file_path, sheet_name='Fast Learners', header=0)

    df_combined_sheet2 = pd.concat([df_combined_sheet2, df_sheet2], ignore_index=True)
    df_combined_sheet3 = pd.concat([df_combined_sheet3, df_sheet3], ignore_index=True)

for idx, row in df_combined_sheet2.iterrows():
    sheet2_roll_counts[row['Roll No.']] = sheet2_roll_counts.get(row['Roll No.'], {'Count': 0, 'Name': 'Unknown'})
    sheet2_roll_counts[row['Roll No.']]['Count'] += 1
    sheet2_roll_counts[row['Roll No.']]['Name'] = row['Student Name']

for idx, row in df_combined_sheet3.iterrows():
    sheet3_roll_counts[row['Roll No.']] = sheet3_roll_counts.get(row['Roll No.'], {'Count': 0, 'Name': 'Unknown'})
    sheet3_roll_counts[row['Roll No.']]['Count'] += 1
    sheet3_roll_counts[row['Roll No.']]['Name'] = row['Student Name']

consolidated_sheet2 = sorted(sheet2_roll_counts.items(), key=lambda x: x[1]['Count'], reverse=True)
consolidated_sheet3 = sorted(sheet3_roll_counts.items(), key=lambda x: x[1]['Count'], reverse=True)

df_consolidated_sheet2 = pd.DataFrame([(roll_no, data['Name'], data['Count']) for roll_no, data in consolidated_sheet2], columns=['Roll No.', 'Student Name', 'Count'])
df_consolidated_sheet3 = pd.DataFrame([(roll_no, data['Name'], data['Count']) for roll_no, data in consolidated_sheet3], columns=['Roll No.', 'Student Name', 'Count'])

top_n = int(input("Enter the count of students: "))

df_consolidated_sheet2_count = df_consolidated_sheet2[df_consolidated_sheet2['Count'] >= top_n]
df_consolidated_sheet3_count = df_consolidated_sheet3[df_consolidated_sheet3['Count'] >= top_n]

output_excel_path = './combined_output_new.xlsx'
with pd.ExcelWriter(output_excel_path) as writer:
    df_sheet1_transposed.to_excel(writer, sheet_name='Student Summary', index=False)
    df_consolidated_sheet2_count.to_excel(writer, sheet_name='Slow Learners', index=False)
    df_consolidated_sheet3_count.to_excel(writer, sheet_name='Fast Learners', index=False)

print("Data saved successfully to", output_excel_path)

