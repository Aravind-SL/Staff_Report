import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


folder_path = './files'
output_path = './output'
output_file = '_output.xlsx'
file_list = os.listdir(folder_path)
excel_files = [file for file in file_list if file.endswith('.xlsx') or file.endswith('.xls')]
if not os.path.exists(output_path):
    os.makedirs(output_path)
for file in excel_files:
    file_path = os.path.join(folder_path, file)
    df = pd.read_excel(file_path, skiprows=10, header=0)

    total_students = len(df)
    total_students_appeared = len(df[df['Marks'] != 'Absent'])
    total_absent = total_students - total_students_appeared
    avg_marks = df[df['Marks'] != 'Absent']['Marks'].astype(float).mean()
    less_than_15 = len(df[df['Marks'].astype(float) < 15])
    between_15_and_30 = len(df[(df['Marks'].astype(float) > 15) & (df['Marks'].astype(float) <= 30)])
    more_than_30 = len(df[df['Marks'].astype(float) > 30])

    results = [{
        'Total Students': total_students,
        'Total Students Appeared': total_students_appeared,
        'Total Absent': total_absent,
        'Average Marks': avg_marks,
        'Students Less than 15': less_than_15,
        'Students Between 15 and 30': between_15_and_30,
        'Students More than 30': more_than_30
    }]
    results_df = pd.DataFrame(results)


    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Sheet1'


    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws1.append(r)


    mark_less_than = float(input(f"Enter a mark to filter students who scored less than that mark for file {file}: "))

    filtered_less_than_df = df[df['Marks'].astype(float) < mark_less_than]
    filtered_less_than_df = filtered_less_than_df.sort_values(by='Marks', ascending=True)
    if 'Unnamed: 4' in filtered_less_than_df:
        filtered_less_than_df= filtered_less_than_df.drop(['Unnamed: 4'], axis=1)
    ws2 = wb.create_sheet('Sheet2')
    for r in dataframe_to_rows(filtered_less_than_df, index=False, header=True):
        ws2.append(r)

    mark_more_than = float(input(f"Enter a mark to filter students who scored more than that mark for file {file}: "))

    filtered_more_than_df = df[df['Marks'].astype(float) > mark_more_than]
    filtered_more_than_df = filtered_more_than_df.sort_values(by='Marks', ascending=False)
    if 'Unnamed: 4' in filtered_more_than_df:
        filtered_more_than_df= filtered_more_than_df.drop(['Unnamed: 4'], axis=1)
    ws3 = wb.create_sheet('Sheet3')
    for r in dataframe_to_rows(filtered_more_than_df, index=False, header=True):
        ws3.append(r)


    output_file_path = os.path.join(output_path, file.replace('.xlsx', '') + output_file)
    wb.save(output_file_path)

    print("Results saved successfully to:", output_file_path)
    
#Task 2    
output_folder_path = './output'
output_files = [file for file in os.listdir(output_folder_path) if file.endswith('.xlsx')]

data_sheet1 = []
for file in output_files:
    file_path = os.path.join(output_folder_path, file)
    df = pd.read_excel(file_path, sheet_name='Sheet1', header=0)
    data_sheet1.append([file[:10]] + df.iloc[0].values.tolist())

columns_sheet1 = ['File Name', 'Total Students', 'Total Students Appeared', 'Total Absent', 
                  'Average Marks', 'Students Less than 15', 
                  'Students Between 15 and 30', 'Students More than 30']

df_sheet1 = pd.DataFrame(data_sheet1, columns=columns_sheet1)

sheet2_roll_counts = {}
sheet3_roll_counts = {}

df_combined_sheet2 = pd.DataFrame()
df_combined_sheet3 = pd.DataFrame()

for file in output_files:
    file_path = os.path.join(output_folder_path, file)

    df_sheet2 = pd.read_excel(file_path, sheet_name='Sheet2', header=0)
    df_sheet3 = pd.read_excel(file_path, sheet_name='Sheet3', header=0)

    df_combined_sheet2 = pd.concat([df_combined_sheet2, df_sheet2], ignore_index=True)
    df_combined_sheet3 = pd.concat([df_combined_sheet3, df_sheet3], ignore_index=True)

for idx, row in df_combined_sheet2.iterrows():
    sheet2_roll_counts[row['Roll No.']] = sheet2_roll_counts.get(row['Roll No.'], {'Count': 0, 'Name': 'Unknown'})
    sheet2_roll_counts[row['Roll No.']]['Count'] += 1
    sheet2_roll_counts[row['Roll No.']]['Name'] = row['Student Name']

for idx, row in df_combined_sheet3.iterrows():
    sheet3_roll_counts[row['Roll No.']] = sheet3_roll_counts.get(row['Roll No.'], {'Count': 0, 'Name': 'Unknown'})
    sheet3_roll_counts[row['Roll No.']]['Count'] += 1
    sheet3_roll_counts[row['Roll No.']]['Name'] = row['Student Name']

top5_sheet2 = sorted(sheet2_roll_counts.items(), key=lambda x: x[1]['Count'], reverse=True)
top5_sheet3 = sorted(sheet3_roll_counts.items(), key=lambda x: x[1]['Count'], reverse=True)

df_top5_sheet2 = pd.DataFrame([(roll_no, data['Name'], data['Count']) for roll_no, data in top5_sheet2], columns=['Roll No.', 'Student Name', 'Count'])
df_top5_sheet3 = pd.DataFrame([(roll_no, data['Name'], data['Count']) for roll_no, data in top5_sheet3], columns=['Roll No.', 'Student Name', 'Count'])

output_excel_path = './combined_output_new.xlsx'
with pd.ExcelWriter(output_excel_path) as writer:
    df_sheet1.to_excel(writer, sheet_name='Sheet1', index=False)
    df_top5_sheet2.to_excel(writer, sheet_name='Sheet2', index=False)
    df_top5_sheet3.to_excel(writer, sheet_name='Sheet3', index=False)

print("Data saved successfully to", output_excel_path)

